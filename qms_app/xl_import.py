import openpyxl
from django.shortcuts import render, redirect
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from .models import QMS, Department  # Assuming your QMS model is in the same app
from datetime import datetime
from .decorators import admin_required
from .forms import ExcelUploadForm # Add the new form
import openpyxl # Import openpyxl
import io 
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
# Assuming you have an admin_required decorator or similar for access control
# from .decorators import admin_required 
from openpyxl.comments import Comment

@login_required
@admin_required
def download_qms_template(request):
    """
    Generates and serves a blank Excel template for QMS data upload.
    Includes all current fields: Location, Remarks, and Involved Departments.
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "QMS_Upload_Template"

    # 1. Define Headers exactly as they are expected by the upload logic
    headers = [
        "S.No", 
        "QMS No", 
        "Remarks", 
        "Involved Departments",
        "Initiation date", 
        "Type", 
        "Department", 
        "Description", 
        "Target Date", 
        "Background", 
        "Location", 
        "Status"
    ]
    sheet.append(headers)

    # 2. Apply styling to header row (Blue background, White bold text)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # 3. Set column widths for better usability
    column_widths = {
    'A': 8,    # S.No
    'B': 20,   # QMS No
    'C': 30,   # Remarks
    'D': 35,   # Involved Departments
    'E': 15,   # Initiation date
    'F': 10,   # Type
    'G': 15,   # Department
    'H': 40,   # Description
    'I': 15,   # Target Date
    'J': 40,   # Background
    'K': 25,   # Location
    'L': 12,   # Status
    }

    for col, width in column_widths.items():
        sheet.column_dimensions[col].width = width

    # 4. Add User Guidance (Comments)
    # Mapping indices based on the headers list (1-based for Excel)
    
    # QMS No is mandatory
    sheet['B1'].comment = Comment('Mandatory. This is the unique identifier.', 'System')
    
    # Date Formats
    date_msg = 'Use YYYY-MM-DD format.'
    sheet['C1'].comment = Comment(date_msg, 'System')
    sheet['G1'].comment = Comment(date_msg, 'System')

    # Choices for Type
    type_choices = ", ".join([c[0] for c in QMS.TYPE_CHOICES])
    sheet['D1'].comment = Comment(f'Valid codes: {type_choices}', 'System')

    # Choices for Department
    dept_choices = ", ".join([c[0] for c in QMS.DEPARTMENT_CHOICES])
    sheet['E1'].comment = Comment(f'Valid Primary Dept codes: {dept_choices}', 'System')

    # Choices for Status
    status_choices = ", ".join([c[0] for c in QMS.STATUS_CHOICES])
    sheet['J1'].comment = Comment(f'Valid codes: {status_choices}', 'System')

    # Many-to-Many Guidance (Involved Departments)
    m2m_msg = (
        "Enter multiple department codes separated by commas.\n"
        "Example: QA, QC, PD\n"
        f"Valid codes: {dept_choices}"
    )
    sheet['L1'].comment = Comment(m2m_msg, 'System')

    # 5. Save to memory and serve as a download
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    response = HttpResponse(
        virtual_workbook.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=QMS_Upload_Template.xlsx'
    
    return response




@login_required
@admin_required
def upload_qms_from_excel(request):
    """
    Handles the upload and processing of the QMS Excel file.
    Only QMS No is mandatory. Updates existing records or creates new ones.
    Empty Excel cells do NOT overwrite existing database data.
    """
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['excel_file']
            try:
                workbook = openpyxl.load_workbook(excel_file, data_only=True)
                sheet = workbook.active
                
                # Extract headers
                header = [str(cell.value).strip() if cell.value else "" for cell in sheet[1]]
                
                # Define expected columns (Matching your model)
                # Note: "Involved Departments" is the column for the M2M field
                required_cols = ["QMS No"]
                optional_cols = [
                    "Initiation date", "Description", "Type", "Target Date", 
                    "Department", "Background", "Status", "Location", 
                    "Remarks", "Involved Departments"
                ]
                
                # Check for mandatory QMS No column
                if "QMS No" not in header:
                    messages.error(request, "The Excel file must contain a 'QMS No' column.")
                    return redirect('upload_qms_from_excel')

                # Create a map of column names to indices
                col_map = {col: header.index(col) for col in (required_cols + optional_cols) if col in header}
                
                created_count = 0
                updated_count = 0

                # Iterate over data rows
                for row_index in range(2, sheet.max_row + 1):
                    row_data = [cell.value for cell in sheet[row_index]]
                    
                    # 1. Get Mandatory QMS Number
                    qms_number = row_data[col_map['QMS No']] if 'QMS No' in col_map else None
                    if not qms_number:
                        continue # Skip rows without a QMS number

                    # 2. Fetch or Create the QMS instance
                    qms_obj, created = QMS.objects.get_or_create(qms_number=str(qms_number.upper()).strip())

                    # 3. Helper function to parse dates
                    def parse_date(val):
                        if isinstance(val, datetime):
                            return val.date()
                        if isinstance(val, str) and val.strip():
                            try:
                                return datetime.strptime(val.strip(), '%Y-%m-%d').date()
                            except ValueError:
                                return None
                        return None

                    # 4. Conditionally Update Fields (Only if Excel cell is NOT empty)
                    # This prevents overwriting existing data with nulls
                    
                    if 'Initiation date' in col_map:
                        dt = parse_date(row_data[col_map['Initiation date']])
                        if dt: qms_obj.initiated_date = dt

                    if 'Target Date' in col_map:
                        dt = parse_date(row_data[col_map['Target Date']])
                        if dt: qms_obj.target_date = dt

                    if 'Description' in col_map and row_data[col_map['Description']]:
                        qms_obj.description = str(row_data[col_map['Description']])

                    if 'Type' in col_map and row_data[col_map['Type']]:
                        qms_obj.type = str(row_data[col_map['Type']]).strip()

                    if 'Department' in col_map and row_data[col_map['Department']]:
                        qms_obj.department = str(row_data[col_map['Department']]).strip()

                    if 'Background' in col_map and row_data[col_map['Background']]:
                        qms_obj.background = str(row_data[col_map['Background']])

                    if 'Status' in col_map and row_data[col_map['Status']]:
                        qms_obj.status = str(row_data[col_map['Status']]).strip()

                    if 'Location' in col_map and row_data[col_map['Location']]:
                        qms_obj.location = str(row_data[col_map['Location']])

                    if 'Remarks' in col_map and row_data[col_map['Remarks']]:
                        qms_obj.remarks = str(row_data[col_map['Remarks']])

                    # Save the basic fields first
                    qms_obj.save()

                    # 5. Handle Many-to-Many Field: Involved Departments
                    # Expects comma separated string: "QA, QC, PD"
                    if 'Involved Departments' in col_map:
                        involved_data = row_data[col_map['Involved Departments']]
                        # Only update if the Excel cell is NOT empty (Partial Update logic)
                        if involved_data and str(involved_data).strip():
                            # Clean the data: "PD, QA, QC" -> ['PD', 'QA', 'QC']
                            # We use .upper() to ensure 'qa' matches 'QA'
                            dept_codes = [
                                code.strip().upper() 
                                for code in str(involved_data).split(',') 
                                if code.strip()
                            ]
                            
                            # Look up Department objects where the 'code' field matches
                            dept_qs = Department.objects.filter(code__in=dept_codes)
                            
                            if dept_qs.exists():
                                # .set() replaces existing relations with the new list
                                qms_obj.involved_departments.set(dept_qs)

                    if created:
                        created_count += 1
                    else:
                        updated_count += 1

                messages.success(request, f"Process complete. Created: {created_count}, Updated: {updated_count}")
                return redirect('qms_list')

            except Exception as e:
                messages.error(request, f"An error occurred: {str(e)}")
                return redirect('upload_qms_from_excel')
    else:
        form = ExcelUploadForm()

    return render(request, 'qms/upload_excel.html', {'form': form})