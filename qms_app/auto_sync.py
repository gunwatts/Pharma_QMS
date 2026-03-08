import os
import openpyxl
from datetime import datetime, timedelta, date # Import date specifically
from pathlib import Path
import logging
import re
from django.conf import settings
from .extenions_code import bulk_create_qms
# --- Configuration (These will be overridden by Django settings when called from views.py) ---
_DEFAULT_QMS_LOGS_DIR = r'Z:\2025 QMS Logs'  # Default directory for QMS Excel files
_DEFAULT_LAST_SYNC_DATE_FILE = Path(__file__).parent / 'qms_last_sync_date.txt'
_DEFAULT_SYNC_ERROR_LOG_FILE = Path(__file__).parent / 'qms_sync_errors.log'
_DEFAULT_SYNC_INTERVAL_HOURS = 24

# --- Logger Setup for this specific module ---
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

if not any(isinstance(handler, logging.FileHandler) and handler.baseFilename == str(_DEFAULT_SYNC_ERROR_LOG_FILE.resolve()) for handler in logger.handlers):
    file_handler = logging.FileHandler(_DEFAULT_SYNC_ERROR_LOG_FILE)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    console_handler = logging.StreamHandler()
    console_handler.setFormatter(formatter)
    logger.addHandler(console_handler)


def _get_last_sync_date(last_sync_date_file_path: Path):
    """Reads the last sync date from a file."""
    if last_sync_date_file_path.exists():
        with open(last_sync_date_file_path, 'r') as f:
            try:
                date_str = f.read().strip()
                return datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S.%f')
            except ValueError:
                logger.error(f"Invalid date format in {last_sync_date_file_path}. Sync date file will be reset upon next successful sync.")
                return None
    return None

def _update_last_sync_date(last_sync_date_file_path: Path):
    """Writes the current datetime as the last sync date to a file."""
    last_sync_date_file_path.parent.mkdir(parents=True, exist_ok=True)
    with open(last_sync_date_file_path, 'w') as f:
        f.write(datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f'))

def parse_excel_date(date_value):
    """
    Parses various Excel date formats into a Python date object.
    Handles datetime objects, date objects, Excel serial numbers, and common string formats.
    """
    if date_value is None:
        return None
    
    # Case 1: Already a Python datetime object
    if isinstance(date_value, datetime):
        return date_value.date()
    # Case 2: Already a Python date object
    if isinstance(date_value, date):
        return date_value
    # Case 3: Excel serial number (float or int)
    elif isinstance(date_value, (int, float)):
        try:
            # Excel serial date for 1900-01-01 is 1. Dates prior to this, or very small numbers, are often errors.
            # Using 60 as a heuristic to avoid 0, 1, or very small numbers that could be invalid.
            if date_value > 60: 
                return datetime.fromtimestamp((date_value - 25569) * 86400).date()
            else:
                return None # Treat very small numbers as invalid
        except Exception: 
            logger.warning(f"Could not convert Excel numeric date: {date_value}")
            return None
    # Case 4: String representation of a date
    elif isinstance(date_value, str):
        cleaned_date_str = date_value.strip().upper()
        if cleaned_date_str in ('NA', 'N/A', '', 'NONE', 'NULL'): # Handle common non-date strings
            return None
        
        # Define common date formats to try
        # Prioritize formats that match Excel's default string representation of dates
        date_formats = [
            '%Y-%m-%d %H:%M:%S', # e.g., '2025-01-10 00:00:00'
            '%Y-%m-%d',         # e.g., '2025-01-10'
            '%m/%d/%Y',         # e.g., '01/10/2025'
            '%d-%m-%Y',         # e.g., '10-01-2025'
            '%Y/%m/%d',         # e.g., '2025/01/10'
            '%d.%m.%Y',         # e.g., '10.01.2025'
        ]
        
        for fmt in date_formats:
            try:
                # Attempt to parse the full string
                return datetime.strptime(date_value, fmt).date() 
            except ValueError:
                pass # Try next format
        
        # If parsing full string fails, try parsing just the date part (before first space)
        if ' ' in date_value:
            date_part = date_value.split(' ')[0]
            for fmt in date_formats: # Re-use formats for date_part
                try:
                    return datetime.strptime(date_part, fmt).date()
                except ValueError:
                    pass
        
        logger.warning(f"Could not parse date string: '{date_value}'")
        return None
    return None # Return None if unable to parse



def auto_sync_qms_from_excel_core(
    qms_model,
    qms_logs_dir: Path = _DEFAULT_QMS_LOGS_DIR,
    last_sync_date_file: Path = _DEFAULT_LAST_SYNC_DATE_FILE,
    sync_error_log_file: Path = _DEFAULT_SYNC_ERROR_LOG_FILE,
    sync_interval_hours: int = _DEFAULT_SYNC_INTERVAL_HOURS,
    force_sync: bool = False,
    show_messages: bool = True
) -> dict:
    results = {
        'success': False,
        'message': '',
        'created_count': 0,
        'updated_count': 0,
        'errors_logged': False,
        'detailed_errors': [],
        'last_sync_attempt': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    # Reconfigure logger if needed
    if Path(sync_error_log_file).resolve() != Path(_DEFAULT_SYNC_ERROR_LOG_FILE).resolve():
        for handler in list(logger.handlers): logger.removeHandler(handler)
        file_handler = logging.FileHandler(sync_error_log_file)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)
        logger.addHandler(file_handler)
        logger.addHandler(logging.StreamHandler())
        logger.setLevel(logging.INFO)

    # Sync Interval Check
    last_sync = _get_last_sync_date(last_sync_date_file)
    current_time = datetime.now()
    if not force_sync and last_sync:
        if current_time - last_sync < timedelta(hours=sync_interval_hours):
            time_to_next_sync = last_sync + timedelta(hours=sync_interval_hours) - current_time
            results['message'] = f"QMS auto-sync not due. Next sync in: {str(time_to_next_sync).split('.')[0]}"
            results['success'] = True
            return results

    logger.info(f"[{results['last_sync_attempt']}] Starting QMS auto-sync (Bulk Optimized).")

    total_created_count = 0
    total_updated_count = 0
    processed_files = []
    skipped_files = []

    # Optimization: Pre-calculate choice sets for O(1) lookups
    TYPE_CHOICES_PREFIXES = [choice[0] for choice in qms_model.TYPE_CHOICES]
    DEPT_CHOICES_SET = {choice[0] for choice in qms_model.DEPARTMENT_CHOICES}
    
    BASE_COL_MAP = {
        '❶': 'initiated_date', '❷': 'qms_number', '❸': 'department',
        '❹': 'description', '❺': 'target_date', '❻': 'background',
    }
    STATUS_COL_NAME = 'Status'

    try:
        qms_logs_dir = Path(qms_logs_dir)
        if not qms_logs_dir.is_dir():
            raise FileNotFoundError(f"QMS logs directory not found: {qms_logs_dir}")

        excel_files = [f for f in os.listdir(qms_logs_dir) if f.lower().endswith(('.xlsx', '.xls', '.xlsm'))]

        for filename in excel_files:
            file_path = qms_logs_dir / filename
            fn_upper = filename.upper()
            
            # 1. Feature: Filename-based Department Override (OOT, OOS, LIR)
            is_qc_override = any(keyword in fn_upper for keyword in settings.DEFAULT_QC_QMS_DEPARTMENT_KEYWORDS)
            
            # 2. Identify QMS Type from filename
            qms_type_found = None
            fn_lower = filename.lower()
            for prefix in TYPE_CHOICES_PREFIXES:
                if re.match(r"^" + re.escape(prefix.lower()) + r"[\s-]?.*", fn_lower):
                    qms_type_found = prefix
                    break
            
            if not qms_type_found:
                skipped_files.append(filename)
                logger.info(f"[{results['last_sync_attempt']}] Skipped '{filename}': Unknown type.")
                continue

            processed_files.append(filename)
            
            try:
                # 3. Fast Loading: Read_only=True
                workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                sheet = workbook.active
                
                header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                header_values = [str(v).strip() if v is not None else '' for v in header_row]
                
                # Dynamic Column Mapping: If QC Override, don't require '❸' (Department)
                current_file_map = BASE_COL_MAP.copy()
                if is_qc_override:
                    current_file_map.pop('❸', None)

                col_indices = {}
                missing_cols = []
                for char, field in current_file_map.items():
                    idx = next((i for i, h in enumerate(header_values) if char in h), None)
                    if idx is not None: col_indices[field] = idx
                    else: missing_cols.append(f"{char}({field})")

                if missing_cols:
                    results['detailed_errors'].append(f"File {filename}: Missing columns {', '.join(missing_cols)}")
                    continue

                status_idx = next((i for i, h in enumerate(header_values) if h and h.strip().lower() == STATUS_COL_NAME.strip().lower()), -1)

                # 4. Process Rows into Memory
                rows_to_process = {} 
                for r_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    qms_num_raw = row[col_indices['qms_number']]
                    if not qms_num_raw or str(qms_num_raw).strip() == '': break
                    
                    qms_number = str(qms_num_raw).strip()
                    
                    try:
                        # Logic for Department (Override vs Column)
                        dept = 'QC' if is_qc_override else str(row[col_indices['department']]).strip()
                        
                        init_date = parse_excel_date(row[col_indices['initiated_date']])
                        targ_date = parse_excel_date(row[col_indices['target_date']])
                        
                        if not init_date: continue
                        if not targ_date: targ_date = init_date + timedelta(days=5)
                        if dept not in DEPT_CHOICES_SET: continue
                        
                        # Status parsing
                        status = 'Open'
                        if status_idx != -1:
                            s_raw = str(row[status_idx]).strip().lower() if row[status_idx] else ''
                            if s_raw in ['closed', 'cancelled', 'canceled']: status = 'Closed'
                            elif s_raw == 'open': status = 'Open'
                            elif s_raw: status = s_raw.upper()

                        rows_to_process[qms_number] = {
                            'initiated_date': init_date,
                            'description': row[col_indices['description']] or '',
                            'type': qms_type_found,
                            'target_date': targ_date,
                            'department': dept,
                            'background': row[col_indices['background']] or '',
                            'status': status,
                        }
                    except Exception: continue

                # 5. Execute Bulk DB Operations for this file
                if rows_to_process:
                    existing_objs = qms_model.objects.filter(qms_number__in=rows_to_process.keys())
                    existing_map = {obj.qms_number: obj for obj in existing_objs}
                    
                    to_create, to_update = [], []
                    for qnum, data in rows_to_process.items():
                        if qnum in existing_map:
                            obj = existing_map[qnum]
                            for field, value in data.items(): setattr(obj, field, value)
                            to_update.append(obj)
                        else:
                            to_create.append(qms_model(qms_number=qnum, **data))

                    if to_create:
                        bulk_create_qms(to_create)
                    if to_update:
                        qms_model.objects.bulk_update(to_update, 
                            ['initiated_date', 'description', 'type', 'target_date', 'department', 'background', 'status'], 
                            batch_size=500)

                    total_created_count += len(to_create)
                    total_updated_count += len(to_update)

                workbook.close()
            except Exception as e:
                results['detailed_errors'].append(f"Error processing {filename}: {str(e)}")
                continue

        # Final Summary
        results.update({
            'created_count': total_created_count,
            'updated_count': total_updated_count,
            'success': True,
            'errors_logged': bool(results['detailed_errors'])
        })
        
        final_msg = f"Auto-sync complete! Created: {total_created_count}, Updated: {total_updated_count}."
        if processed_files: final_msg += f"\nProcessed: {', '.join(processed_files)}"
        if skipped_files: final_msg += f"\nSkipped: {', '.join(skipped_files)}"
        if show_messages:
            results['message'] = final_msg
        _update_last_sync_date(last_sync_date_file)
        logger.info(f"[{results['last_sync_attempt']}] {final_msg}")

    except Exception as e:
        results['message'] = f"Fatal error during sync: {str(e)}"
        logger.exception(f"[{results['last_sync_attempt']}] {results['message']}")

    return results