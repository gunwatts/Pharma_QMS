from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from .models import QMS, ActionPlan, MiniQMS, Note, MiniAction, UpdateRequest, Department
from .forms import QMSForm, ActionPlanForm, MiniQMSForm, NoteForm, MiniActionForm, UserCreationForm, UpdateRequestForm
from django.contrib.auth.models import User, Group
from django.db.models import Q, F
from datetime import date, timedelta
from .decorators import admin_required
from django.core.exceptions import PermissionDenied
from collections import Counter # Import Counter
import json # Import json
from django.http import FileResponse, Http404
from django.views.decorators.csrf import csrf_exempt
import openpyxl # Import openpyxl
import io 
from django.http import HttpResponse
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import date # Ensure datetime is imported
import re
import os
from django.shortcuts import redirect
from .auto_sync import _get_last_sync_date, auto_sync_qms_from_excel_core
from django.conf import settings
from datetime import datetime
from django.db.models import Count
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from .serializers import QMSSerializer, ActionPlanSerializer, DepartmentSerializer

# --- Dashboard View (Updated for Chart & Layout) ---

@login_required
def dashboard(request):
    return render(request, 'qms/loading.html')
@login_required
def dashboard_final(request):
    """
    This function prepares all data for the main dashboard, including context for
    the interactive chart and dynamic tables.
    """
    QMS_LOGS_DIRS = settings.QMS_LOGS_DIRS
    last_sync = _get_last_sync_date(settings.QMS_LAST_SYNC_DATE_FILE)
    current_time = datetime.now()
    sync_interval = timedelta(hours=settings.QMS_SYNC_INTERVAL_HOURS)
    if (not settings.QMS_LAST_SYNC_DATE_FILE.exists() ) or current_time - last_sync >= sync_interval:
        for i, current_path in enumerate(QMS_LOGS_DIRS):
            sync_results = auto_sync_qms_from_excel_core(
                qms_model=QMS, # Pass the QMS model class
                qms_logs_dir=current_path,
                last_sync_date_file=settings.QMS_LAST_SYNC_DATE_FILE,
                sync_error_log_file=settings.QMS_SYNC_ERROR_LOG_FILE,
                sync_interval_hours=settings.QMS_SYNC_INTERVAL_HOURS
            )

            if not sync_results['success']:
                messages.error(request, f"QMS Auto-sync encountered an issue: {sync_results['message']}. Please check logs.")
                if sync_results['detailed_errors']:
                    for error in sync_results['detailed_errors']:
                        messages.warning(request, f"Sync detail: {error}")
            elif sync_results['created_count'] > 0 or sync_results['updated_count'] > 0:
                print_path= current_path[3:]
                messages.success(request, f"Path {print_path}: {sync_results['message']}")
            # else:
            #     # If sync was skipped because not due
            #     messages.info(request, f"QMS Auto-sync: {sync_results['message']}")

            # Re-read last sync date for display if it was updated
            last_sync_display = "Never synced"
            if settings.QMS_LAST_SYNC_DATE_FILE.exists():
                with open(settings.QMS_LAST_SYNC_DATE_FILE, 'r') as f:
                    try:
                        last_sync_raw = f.read().strip()
                        last_sync_dt = datetime.strptime(last_sync_raw, '%Y-%m-%d %H:%M:%S.%f')
                        last_sync_display = last_sync_dt.strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        last_sync_display = "Error reading last sync date"


            if (i < len(QMS_LOGS_DIRS) - 1) and settings.QMS_LAST_SYNC_DATE_FILE.exists() :
            # Delete the sync date file so the next iteration doesn't think it's "too early" to sync
                os.remove(settings.QMS_LAST_SYNC_DATE_FILE)



    today = date.today()
    seven_days_later = today + timedelta(days=7)
    
    # This query is the source of the 'near_due_qms' data. It must be a clean QuerySet.
    near_due_qms = QMS.objects.filter(
        target_date__gte=today,
        target_date__lte=seven_days_later,
        status='Open'
    )
    
    # Data for the main filterable table
    action_plans = ActionPlan.objects.filter(status='Open')
    mini_qms = MiniQMS.objects.filter(status='Open')
    
    # This logic correctly builds the combined list for the main table
    combined_list = []
    for item in action_plans:
        combined_list.append({
            'source': 'Action Plan', 'obj': item, 'qms_number': item.qms.qms_number,
            'description': item.action_plan, 'target_date': item.target_date,
            'department': item.get_department_display(),
        })
    for item in mini_qms:
        combined_list.append({
            'source': 'Mini QMS', 'obj': item, 'qms_number': item.qms_number,
            'description': item.description, 'target_date': item.proposed_target_date,
            'department': item.get_department_display(),
        })
    
    combined_list.sort(key=lambda x: x['target_date'])

    # Data for the pie chart
    department_counts = Counter()
    all_open_actions = ActionPlan.objects.filter(status='Open')
    all_open_mini_qms = MiniQMS.objects.filter(status='Open')
    
    for action in all_open_actions:
        department_counts[action.get_department_display()] += 1
    for mini_qms_item in all_open_mini_qms:
        department_counts[mini_qms_item.get_department_display()] += 1

    top_4_departments = department_counts.most_common(4)
    chart_labels = json.dumps([dept[0] for dept in top_4_departments])
    chart_data = json.dumps([dept[1] for dept in top_4_departments])

    # Data for analytics cards
    qms_by_type = {choice[1]: QMS.objects.filter(type=choice[0]).count() for choice in QMS.TYPE_CHOICES}
    open_actions_count = all_open_actions.count()
    closed_actions_count = ActionPlan.objects.filter(status='Closed').count()
    total_actions_count = open_actions_count + closed_actions_count
    closed_percentage = (closed_actions_count / total_actions_count * 100) if total_actions_count > 0 else 0

    # Data for JS interactivity
    dept_map = {name: code for code, name in QMS.DEPARTMENT_CHOICES}

    context = {
        'near_due_qms': near_due_qms,
        'combined_list': combined_list,
        'department_choices': QMS.DEPARTMENT_CHOICES,
        'qms_by_type': qms_by_type,
        'open_actions': open_actions_count,
        'closed_actions': closed_actions_count,
        'closed_percentage': round(closed_percentage, 2),
        'chart_labels': chart_labels,
        'chart_data': chart_data,
        'department_map_json': json.dumps(dept_map),
    }
    return render(request, 'qms/dashboard.html', context)

# --- QMS CRUD Views (Updated list view) ---
@login_required
def qms_list(request):
    qms_list_data = QMS.objects.exclude(status='Closed')

    # Get all department choices for the dropdown
    # We use QMS.DEPARTMENT_CHOICES directly to ensure all possible options are listed.
    qms_departments = QMS.DEPARTMENT_CHOICES

    # Aggregate QMS count by department for the chart
    department_counts_query = QMS.objects.values('department').annotate(count=Count('department'))

    qms_data = {}
    # Initialize all possible departments with a count of 0
    # This ensures all departments appear on the chart, even if no QMS are linked,
    # and provides consistent labels.
    all_department_names_map = dict(QMS.DEPARTMENT_CHOICES)
    for choice_code, choice_name in QMS.DEPARTMENT_CHOICES:
        qms_data[choice_name] = 0

    # Populate the counts for departments that actually have QMS entries
    for entry in department_counts_query:
        # Get the display name for the department code
        department_display_name = all_department_names_map.get(entry['department'])
        if department_display_name:
            qms_data[department_display_name] = entry['count']

    # Convert the Python dict to a JSON string for the template
    # `json.dumps` handles the correct formatting. `|safe` in the template
    # is necessary because we trust this content and don't want Django to escape it.
    qms_data_json = json.dumps(qms_data)

    context = {
        'qms_list': qms_list_data,
        'qms_departments': qms_departments, # Pass all department choices for dropdown
        'qms_data_json': qms_data_json,     # Pass chart data as JSON
    }
    return render(request, 'qms/qms_list.html', context)
# --- Mini QMS CRUD Views (Updated list view) ---

@admin_required
@login_required
def sync_excel_data(request):
    file_path = r"C:\Users\GanapathiPolisetti\Visual code\QMS_Tracker\qms_tracker\qms_last_sync_date.txt"
    # Check if file exists
    if os.path.exists(file_path):
        os.remove(file_path)
    return redirect('dashboard')



@login_required
def mini_qms_list(request):
    # Only show 'Open' Mini QMS records
    mini_qms_list = MiniQMS.objects.filter(status='Open')
    return render(request, 'qms/mini_qms_list.html', {'mini_qms_list': mini_qms_list})

# --- New Closure Views ---
@login_required
@admin_required
def qms_close(request, pk):
    if request.method == 'POST':
        qms = get_object_or_404(QMS, pk=pk)
        # Close all related Action Plans first
        qms.action_plans.update(status='Closed')
        # Then close the QMS itself
        qms.status = 'Closed'
        qms.save()
        messages.success(request, f"QMS '{qms.qms_number}' and all related action plans have been closed.")
    return redirect('qms_detail', pk=pk)

@login_required
@admin_required
def mini_qms_close(request, pk):
    if request.method == 'POST':
        mini_qms = get_object_or_404(MiniQMS, pk=pk)
        mini_qms.status = 'Closed'
        mini_qms.save()
        messages.success(request, f"Mini QMS '{mini_qms.qms_number}' has been closed.")
    return redirect('mini_qms_list')

# --- New Closed List View ---
@login_required
def closed_list(request):
    closed_qms = QMS.objects.filter(status='Closed')
    closed_mini_qms = MiniQMS.objects.filter(status='Closed')
    context = {
        'closed_qms': closed_qms,
        'closed_mini_qms': closed_mini_qms,
    }
    return render(request, 'qms/closed_list.html', context)


@login_required
@admin_required
def admin_panel(request):
    if request.method == 'POST':
        if 'create_user' in request.POST:
            form = UserCreationForm(request.POST)
            if form.is_valid():
                user = form.save(commit=False)
                user.set_password('User@1234')
                user.save()
                if form.cleaned_data.get('is_admin'):
                    admin_group = Group.objects.get(name='admin')
                    user.groups.add(admin_group)
                messages.success(request, f"User '{user.username}' created successfully.")
                return redirect('admin_panel')
        elif 'reset_password' in request.POST:
            user_id = request.POST.get('user_id')
            user_to_reset = get_object_or_404(User, id=user_id)
            user_to_reset.set_password('User@1234')
            user_to_reset.save()
            messages.info(request, f"Password for '{user_to_reset.username}' has been reset.")
            return redirect('admin_panel')
    form = UserCreationForm()
    users = User.objects.all().prefetch_related('groups')
    return render(request, 'qms/admin_panel.html', {'form': form, 'users': users})


@login_required
def qms_detail(request, pk):
    qms = get_object_or_404(QMS, pk=pk)
    action_form = ActionPlanForm()
    return render(request, 'qms/qms_detail.html', {'qms': qms, 'action_form': action_form})

@login_required
@admin_required
def qms_create(request):
    if request.method == 'POST':
        form = QMSForm(request.POST,request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'QMS created successfully.')
            return redirect('qms_list')
    else:
        form = QMSForm()
    return render(request, 'qms/qms_form.html', {'form': form, 'form_title': 'Create New QMS'})

@login_required
@admin_required
def qms_edit(request, pk):
    qms = get_object_or_404(QMS, pk=pk)
    if request.method == 'POST':
        form = QMSForm(request.POST,request.FILES, instance=qms)
        if form.is_valid():
            form.save()
            messages.success(request, 'QMS updated successfully.')
            return redirect('qms_detail', pk=qms.pk)
    else:
        form = QMSForm(instance=qms)
    return render(request, 'qms/qms_form.html', {'form': form, 'form_title': 'Edit QMS'})

@login_required
@admin_required
def qms_delete(request, pk):
    qms = get_object_or_404(QMS, pk=pk)
    if request.method == 'POST':
        if qms.pdf_file:
            qms.pdf_file.delete(save=False)  # Delete the associated PDF file
        qms.delete()
        messages.success(request, 'QMS deleted successfully.')
        return redirect('qms_list')
    return render(request, 'qms/confirm_delete.html', {'object': qms})

# --- Action Plan CRUD Views ---

@login_required
@admin_required
def action_plan_create(request, qms_pk):
    qms = get_object_or_404(QMS, pk=qms_pk)
    if request.method == 'POST':
        form = ActionPlanForm(request.POST)
        if form.is_valid():
            action_plan = form.save(commit=False)
            action_plan.qms = qms
            action_plan.save()
            messages.success(request, 'Action Plan added successfully.')
    return redirect('qms_detail', pk=qms_pk)

@login_required
@admin_required
def action_plan_edit(request, pk):
    action_plan = get_object_or_404(ActionPlan, pk=pk)
    if request.method == 'POST':
        form = ActionPlanForm(request.POST, instance=action_plan)
        if form.is_valid():
            form.save()
            messages.success(request, 'Action Plan updated successfully.')
            return redirect('qms_detail', pk=action_plan.qms.pk)
    else:
        form = ActionPlanForm(instance=action_plan)
    return render(request, 'qms/action_plan_form.html', {'form': form, 'action_plan': action_plan})

@login_required
@admin_required
def action_plan_delete(request, pk):
    action_plan = get_object_or_404(ActionPlan, pk=pk)
    qms_pk = action_plan.qms.pk
    if request.method == 'POST':
        action_plan.delete()
        messages.success(request, 'Action Plan deleted successfully.')
        return redirect('qms_detail', pk=qms_pk)
    return render(request, 'qms/confirm_delete.html', {'object': action_plan})

# --- Mini QMS CRUD Views ---

@login_required
@admin_required
def mini_qms_create(request):
    if request.method == 'POST':
        form = MiniQMSForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mini QMS created successfully.')
            return redirect('mini_qms_list')
    else:
        form = MiniQMSForm()
    return render(request, 'qms/mini_qms_form.html', {'form': form, 'form_title': 'Create New Mini QMS'})

@login_required
@admin_required
def mini_qms_edit(request, pk):
    mini_qms = get_object_or_404(MiniQMS, pk=pk)
    if request.method == 'POST':
        form = MiniQMSForm(request.POST, instance=mini_qms)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mini QMS updated successfully.')
            return redirect('mini_qms_list')
    else:
        form = MiniQMSForm(instance=mini_qms)
    return render(request, 'qms/mini_qms_form.html', {'form': form, 'form_title': 'Edit Mini QMS'})

@login_required
@admin_required
def mini_qms_delete(request, pk):
    mini_qms = get_object_or_404(MiniQMS, pk=pk)
    if request.method == 'POST':
        mini_qms.delete()
        messages.success(request, 'Mini QMS deleted successfully.')
        return redirect('mini_qms_list')
    return render(request, 'qms/confirm_delete.html', {'object': mini_qms})

# --- Note CRUD Views (Owner only) ---

@login_required
def notes_list(request):
    notes = Note.objects.filter(owner=request.user)
    if request.method == 'POST':
        form = NoteForm(request.POST)
        if form.is_valid():
            note = form.save(commit=False)
            note.owner = request.user
            note.save()
            messages.success(request, "Note created successfully.")
            return redirect('notes_list')
    else:
        form = NoteForm()
    return render(request, 'qms/notes_list.html', {'notes': notes, 'form': form})

@login_required
def note_edit(request, pk):
    note = get_object_or_404(Note, pk=pk)
    if note.owner != request.user:
        raise PermissionDenied
    if request.method == 'POST':
        form = NoteForm(request.POST, instance=note)
        if form.is_valid():
            form.save()
            messages.success(request, 'Note updated successfully.')
            return redirect('notes_list')
    else:
        form = NoteForm(instance=note)
    return render(request, 'qms/note_form.html', {'form': form})

@login_required
def note_delete(request, pk):
    note = get_object_or_404(Note, pk=pk)
    if note.owner != request.user:
        raise PermissionDenied
    if request.method == 'POST':
        note.delete()
        messages.success(request, 'Note deleted successfully.')
        return redirect('notes_list')
    return render(request, 'qms/confirm_delete.html', {'object': note})

# --- Mini Action CRUD Views (Owner only) ---

@login_required
def mini_actions_list(request):
    actions = MiniAction.objects.filter(owner=request.user).order_by('target_date')
    if request.method == 'POST':
        form = MiniActionForm(request.POST)
        if form.is_valid():
            action = form.save(commit=False)
            action.owner = request.user
            action.save()
            messages.success(request, "Mini Action created successfully.")
            return redirect('mini_actions_list')
    else:
        form = MiniActionForm()
    return render(request, 'qms/mini_actions_list.html', {'actions': actions, 'form': form})

@login_required
def mini_action_edit(request, pk):
    action = get_object_or_404(MiniAction, pk=pk)
    if action.owner != request.user:
        raise PermissionDenied
    if request.method == 'POST':
        form = MiniActionForm(request.POST, instance=action)
        if form.is_valid():
            form.save()
            messages.success(request, 'Mini Action updated successfully.')
            return redirect('mini_actions_list')
    else:
        form = MiniActionForm(instance=action)
    return render(request, 'qms/mini_action_form.html', {'form': form})

@login_required
def mini_action_delete(request, pk):
    action = get_object_or_404(MiniAction, pk=pk)
    if action.owner != request.user:
        raise PermissionDenied
    if request.method == 'POST':
        action.delete()
        messages.success(request, 'Mini Action deleted successfully.')
        return redirect('mini_actions_list')
    return render(request, 'qms/confirm_delete.html', {'object': action})


def download_pbix(request):
    file_path = r"Z:\QMS Analytics.pbix"  # Full file path
    if os.path.exists(file_path):
        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename="QMS Analytics.pbix")
    else:
        raise Http404("File not found.")

@login_required
def qms_pdf_view(request, pk):
    """
    A dedicated view to display a PDF file in a clean template.
    This avoids iframe security issues.
    """
    qms = get_object_or_404(QMS, pk=pk)
    return render(request, 'qms/pdf_viewer.html', {'qms': qms})


# --- New Update Request Views ---
@login_required
def submit_update_request(request):
    if request.method == 'POST':
        form = UpdateRequestForm(request.POST, request.FILES)
        if form.is_valid():
            update_request = form.save(commit=False)
            update_request.owner = request.user
            update_request.save()
            messages.success(request, 'Your update request has been submitted successfully.')
            return redirect('dashboard')
    else:
        form = UpdateRequestForm()
    return render(request, 'qms/update_request_form.html', {'form': form})

@login_required
@admin_required
def list_update_requests(request):
    requests = UpdateRequest.objects.all()
    return render(request, 'qms/update_request_list.html', {'requests': requests})

@login_required
@admin_required
def delete_update_request(request, pk):
    req = get_object_or_404(UpdateRequest, pk=pk)
    if request.method == 'POST':
        req.delete()
        messages.success(request, f"Request for QMS '{req.qms_number}' has been deleted.")
        return redirect('list_update_requests')
    # This view is POST only
    


@csrf_exempt
def export_qms_to_excel(request,dept=None):
    if request.method != 'POST':
        return render(request, 'qms/export_confirm.html', {'dept': dept})
    # Auto-sync logic before exporting
    if settings.QMS_LAST_SYNC_DATE_FILE.exists():
        os.remove(settings.QMS_LAST_SYNC_DATE_FILE)
    QMS_LOGS_DIRS = settings.QMS_LOGS_DIRS
    last_sync = _get_last_sync_date(settings.QMS_LAST_SYNC_DATE_FILE)
    current_time = datetime.now()
    sync_interval = timedelta(hours=settings.QMS_SYNC_INTERVAL_HOURS)
    if (not settings.QMS_LAST_SYNC_DATE_FILE.exists() ) or current_time - last_sync >= sync_interval:
        for i, current_path in enumerate(QMS_LOGS_DIRS):
            sync_results = auto_sync_qms_from_excel_core(
                qms_model=QMS, # Pass the QMS model class
                qms_logs_dir=current_path,
                last_sync_date_file=settings.QMS_LAST_SYNC_DATE_FILE,
                sync_error_log_file=settings.QMS_SYNC_ERROR_LOG_FILE,
                sync_interval_hours=settings.QMS_SYNC_INTERVAL_HOURS,
                show_messages=False  # Suppress messages during export sync
            )


            last_sync_display = "Never synced"
            if settings.QMS_LAST_SYNC_DATE_FILE.exists():
                with open(settings.QMS_LAST_SYNC_DATE_FILE, 'r') as f:
                    try:
                        last_sync_raw = f.read().strip()
                        last_sync_dt = datetime.strptime(last_sync_raw, '%Y-%m-%d %H:%M:%S.%f')
                        last_sync_display = last_sync_dt.strftime('%Y-%m-%d %H:%M:%S')
                    except ValueError:
                        last_sync_display = "Error reading last sync date"


            if (i < len(QMS_LOGS_DIRS) - 1) and settings.QMS_LAST_SYNC_DATE_FILE.exists() :
            # Delete the sync date file so the next iteration doesn't think it's "too early" to sync
                os.remove(settings.QMS_LAST_SYNC_DATE_FILE)




    valid_depts = dict(QMS.DEPARTMENT_CHOICES).keys()

    all_open_qms = QMS.objects.exclude(status='Closed')


    

    if dept and dept.upper() in valid_depts:
        all_open_qms_involvd = all_open_qms.exclude(
            department=dept.upper()
        ).filter(
            involved_departments__code=dept.upper()
        ).distinct()
        all_open_qms = all_open_qms.filter(department=dept.upper())
        name_prefix = f"{dept.upper()}_"
    else:
        name_prefix = ""
        all_open_qms_involvd =None

    all_open_qms = all_open_qms.order_by('target_date')
    # 1. Fetch Data
    
    today = timezone.now().date()

    # 2. Setup Workbook and Styles
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active) # Remove default sheet

    def write_qms_sheet(wb, sheet_name, qms_list):
        """Helper function to create a sheet and fill it with QMS data"""
        if not qms_list:
            return # Skip creating sheet if no data

        sheet = wb.create_sheet(title=sheet_name[:31]) # Excel limits title to 31 chars
        
        # Headers (Column H is Remarks, Column I is Due Status, Column J is Status)
        headers = [
            "S.No.", "QMS No", "Initiated Date", "Description", 
            "Target Date", "Dept", "Type", "Background", "Remarks", "Involved Departments", "Due Status", "Status"
        ]
        sheet.append(headers)

        # Header Styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Write Data
        for idx, qms in enumerate(qms_list, start=1):
            target_date_cell = f"E{idx + 1}"
            # Formula checks if Target Date < TODAY()
            due_status_formula = f'=IF({target_date_cell}<TODAY(), "Overdue", "")'
            
            row_data = [
                idx,
                str(qms.qms_number),
                qms.initiated_date,
                str(qms.description or ""),
                qms.target_date,
                str(qms.department or ""),
                str(qms.type),
                str(qms.background or ""),
                str(qms.remarks or ""),  # NEW COLUMN
                qms.responsibilities(),  # NEW COLUMN
                due_status_formula,
                qms.get_status_display() if hasattr(qms, 'get_status_display') else str(qms.status)
            ]
            sheet.append(row_data)

        # Table and Cell Formatting
        max_row = sheet.max_row
        max_col_letter = openpyxl.utils.get_column_letter(sheet.max_column)
        
        # Add Table
        sanitized_name = re.sub(r'[^a-zA-Z0-9]', '', sheet_name)
        table = Table(displayName=f"Table_{sanitized_name}", ref=f"A1:{max_col_letter}{max_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        sheet.add_table(table)

        # Style data cells
        for row in sheet.iter_rows(min_row=2, max_row=max_row):
            for cell in row:
                cell.alignment = center_align
                cell.border = thin_border
            row[2].number_format = 'dd mmm yyyy' # Initiated Date (C)
            row[4].number_format = 'dd mmm yyyy' # Target Date (E)

        # Auto-fit columns
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 3)
            # Description (E) and 
            if column in ['D']: 
                sheet.column_dimensions[column].width = min(adjusted_width, 60)
            elif column in ['H']: # Background can be longer, but still limit to prevent extreme widths
                sheet.column_dimensions[column].width = min(adjusted_width, 28)
            elif column in ["I"]:
                sheet.column_dimensions[column].width = min(adjusted_width, 40)
            elif column in ["J"]:
                sheet.column_dimensions[column].width = min(adjusted_width, 13)
            else:
                sheet.column_dimensions[column].width = adjusted_width

    # --- SHEET 1: Overdue ---
    overdue_qms = [q for q in all_open_qms if q.target_date and q.target_date < today and q.status == 'Open']
    write_qms_sheet(workbook, "Overdue", overdue_qms)

    # --- SHEET 2: Due Soon (Next 7 Days) ---
    due_soon_qms = [
        q for q in all_open_qms
        if q.target_date and today <= q.target_date <= today + timezone.timedelta(days=7)
        and q.status == 'Open']
    write_qms_sheet(workbook, "Due Soon", due_soon_qms)

    #invloved qms sheet only if dept filter is applied and there are involved qms
    if all_open_qms_involvd is not None:
        write_qms_sheet(workbook, "Related Dept", all_open_qms_involvd)

    # --- SHEET 3  ---
    all_open_sorted_initiation = sorted(all_open_qms, key=lambda x: x.initiated_date)
    cft_qms = [q for q in all_open_sorted_initiation if q.status == 'CFT']
    write_qms_sheet(workbook, "Under CFT review", cft_qms)

    # --- SHEET 4: All Open (Sorted by Target Date)---
    all_open_sorted = sorted(all_open_qms, key=lambda x: x.target_date if x.target_date else today)
    all_open_qms_only = [q for q in all_open_sorted if q.status == 'Open'] # Filter to only 'Open' for the main sheet
    write_qms_sheet(workbook, "All Open", all_open_qms_only)

    # --- SHEET 5: EM ---
    em_qms = [q for q in all_open_qms if q.status == 'EM']
    write_qms_sheet(workbook, "Effectiveness Monitoring", em_qms)

    # --- TYPE SPECIFIC SHEETS ---
    # Prepare sorting logic for QMS Number
    def get_qms_sort_key(qms):
        match = re.search(r'(\d{2,4})-(\d{3})', str(qms.qms_number))
        if match:
            year = int("20" + match.group(1)) if len(match.group(1)) == 2 else int(match.group(1))
            number = int(match.group(2))
            return (year, number)
        return (0, 0)

    # Group by type
    qms_by_type = {}
    for qms in all_open_qms:
        t_name = qms.get_type_display()
        if t_name not in qms_by_type: qms_by_type[t_name] = []
        qms_by_type[t_name].append(qms)

    # Write each type sheet
    for type_code, type_name in QMS.TYPE_CHOICES:
        if type_name in qms_by_type:
            sorted_list = sorted(qms_by_type[type_name], key=get_qms_sort_key)
            write_qms_sheet(workbook, type_name, sorted_list)

    # Finalize Response
    ws = workbook.active or workbook.create_sheet("QMS")
    if not all_open_qms.exists():
        ws.append([f"Nil QMS for {dept.upper()} Department"])
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    response = HttpResponse(
        virtual_workbook.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={ name_prefix}Detailed_QMS_Report.xlsx'
    return response




def export_department_qms_to_excel(request):
    today = date.today()
    date_filter = request.POST.get("date_filter", "all")
    if date_filter == "7days":
        timeline = today + timedelta(days=7)
    elif date_filter == "month":
        timeline = today + timedelta(days=30) 
    else:
        timeline = today + timedelta(days=365*10) # Default to 10 years in the future if "all" is selected

    # Filter QMS instances that are 'Open' and have a target_date within the next 7 days (inclusive of today)
    # Also include any QMS that are overdue (target_date < today) for comprehensive reporting
    department_qms = QMS.objects.filter(
        status='Open'
    ).order_by('department', 'target_date') # Order by department, then by target_date

    # Group QMS by department
    qms_by_department = {}
    for qms in department_qms:
        # Check if the QMS is due within the next 7 days or is overdue
        if qms.target_date <= timeline:
            dept_name = qms.get_department_display()
            if dept_name not in qms_by_department:
                qms_by_department[dept_name] = []
            qms_by_department[dept_name].append(qms)

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active) 

    # --- Define Styles ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    date_style = 'dd mmm yyyy'
    
    # --- Headers ---
    headers = [
        "S.No.", "QMS Number", "Type", "Initiated Date", "Description", 
        "Target Date", "Background", "Remarks", "Due Status"
    ]
    
    sheet_index = 0
    # Iterate through departments that have relevant QMS
    for dept_code, dept_name in QMS.DEPARTMENT_CHOICES:
        if dept_name in qms_by_department:
            # Create a sheet for each department
            sheet = workbook.create_sheet(title=dept_name, index=sheet_index)
            sheet_index += 1
            
            # Write headers
            sheet.append(headers)

            # Apply header styles
            for cell in sheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = thin_border

            # Write data rows
            serial_number = 1
            # Sort QMS within each department by target_date (already done in initial query, but good to ensure)
            sorted_qms_list = sorted(qms_by_department[dept_name], key=lambda x: x.target_date)

            for qms in sorted_qms_list:
                # Target Date is now in column F (6th column)
                target_date_cell = f"F{serial_number + 1}"
                due_status_formula = f'=IF({target_date_cell}<TODAY(), "Overdue", IF({target_date_cell}<=TODAY()+7, "Due Soon", ""))'

                row_data = [
                    serial_number,
                    qms.qms_number,
                    qms.get_type_display(), # Added Type back for department sheets
                    qms.initiated_date,
                    qms.description,
                    qms.target_date,
                    qms.background,
                    qms.remarks,
                    due_status_formula
                ]
                sheet.append(row_data)
                serial_number += 1
            
            # --- Apply Formatting and Create Table ---
            max_row = sheet.max_row
            max_col_letter = openpyxl.utils.get_column_letter(sheet.max_column)
            table_range = f"A1:{max_col_letter}{max_row}"
            
            table = Table(displayName=f"{dept_name.replace(' ', '')}Table", ref=table_range)
            style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                   showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            table.tableStyleInfo = style
            sheet.add_table(table)
            
            # --- Apply styles to all data cells ---
            for row in sheet.iter_rows(min_row=2, max_row=max_row):
                for cell in row:
                    cell.alignment = center_alignment
                    cell.border = thin_border
                
                # Column indexes for date formatting: Initiated Date (D), Target Date (F)
                row[3].number_format = date_style 
                row[5].number_format = date_style 

            # --- Auto-fit column widths ---
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                # Description (E) and Background (G) and Remarks (H) columns
                if column in ['E', 'G', 'H']: 
                    sheet.column_dimensions[column].width = min(adjusted_width, 60)
                else:
                    sheet.column_dimensions[column].width = adjusted_width

    # Create a virtual file in memory
    virtual_workbook = io.BytesIO()
    workbook.save(virtual_workbook)
    virtual_workbook.seek(0)
    
    response = HttpResponse(
        virtual_workbook.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=department_qms_report.xlsx'
    return response




@api_view(['GET'])
@permission_classes([AllowAny])
def qms_api(request):
    qs = QMS.objects.all()
    return Response(QMSSerializer(qs, many=True).data)


@api_view(['GET'])
@permission_classes([AllowAny])
def action_plans_api(request):
    qs = ActionPlan.objects.select_related('qms')
    return Response(ActionPlanSerializer(qs, many=True).data)


@api_view(['GET'])
@permission_classes([AllowAny])
def departments_api(request):
    qs = Department.objects.all()
    return Response(DepartmentSerializer(qs, many=True).data)


@api_view(['GET'])
@permission_classes([AllowAny])
def qms_involved_departments_api(request):
    result = []

    for qms in QMS.objects.all():
        deps = qms.involved_departments.all()
        if deps.exists():
            for d in deps:
                result.append({
                    "qms_id": qms.id,
                    "qms_number": qms.qms_number,
                    "department_code": d.code,
                    "department_name": d.name,
                })
        else:
            result.append({
                "qms_id": qms.id,
                "qms_number": qms.qms_number,
                "department_code": "Nil",
                "department_name": "Nil",
            })

    return Response(result)